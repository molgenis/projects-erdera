"""Build excel template from schema"""

from os import environ, path
import sys
import logging
from typing import TypedDict
import xlsxwriter
from openpyxl.utils.cell import get_column_letter
from molgenis_emx2_pyclient import Client
from molgenis_emx2_pyclient.metadata import Schema, Table, Column
from dotenv import load_dotenv
load_dotenv()

logging.captureWarnings(True)
logging.basicConfig(stream=sys.stdout, level=logging.DEBUG)
log = logging.getLogger("Template Generator")

# set params
SCHEMA: str = "rd3"
TABLES: list[str] = ['Samples OGM', 'Experiments OGM']

# SCHEMA: str = "pet store"
# TABLES: list[str] = ['Pet', 'Order', 'Category']

MAX_TEMPLATE_ROWS: int = 1000

OUTPUT_FILE: str = environ.get('OUTPUT_FILE')

client = Client(url=environ['MOLGENIS_HOST'], token=environ['MOLGENIS_TOKEN'])


class WorkbookStyles(TypedDict):
    """Workbook formats"""
    header_required: dict
    header_default: dict
    cell_required: dict


class BuildTemplate:
    """Build template"""

    def __init__(self,
                 schema: str,
                 tables: list[str],
                 max_template_rows: int = 250,
                 sys_output_filename: str = None):
        """New template generator

        :param schema: name of the schema
        :type schema: str

        :param tables: names of the tables that exist in the schema
        :type tables: str[]

        :param max_template_rows: number of rows to prefill with styles, validation, etc.
        :type max_template_rows: int

        """
        self.output_filename = f"{schema}.xlsx"
        self.schema = schema
        self.tables = tables

        self.max_template_rows = max_template_rows

        self.should_build_lookup_sheet = False
        self.lookups_col_index = 0
        self.lookups = []

        if sys_output_filename:
            output_basename = path.basename(sys_output_filename)
            self.output_filename = sys_output_filename.replace(
                output_basename, self.output_filename)

    def column_is_required(self, column: Column) -> bool:
        """Determine if a column is required based on schema metadata

        :param column: metadata object for a column
        :type column: Column

        :returns: bool
        """
        is_key = column.key > 0 if column.get('key') else False
        is_req = column.get('required')
        return is_key and is_req

    def write_sheet_header(self,
                           sheet,
                           column: Column,
                           styles: WorkbookStyles,
                           col_index: int = 0):
        """Write a column header to a sheet

        :param sheet: a workbook.worksheet

        :param column: metadata object for a column
        :type column: Column

        :param col_index: the column to write the header into (zero index)
        :type col_index: int

        """
        current_header_style = styles['header_default']
        if self.column_is_required(column=column):
            current_header_style = styles['header_required']

        sheet.write(0, col_index, column.name, current_header_style)

    def column_is_ontology_type(self, column: Column) -> bool:
        """Determine if the column is ONTOLOGY or ONTOLOGY_ARRAY"""
        return column.columnType.startswith('ONTOLOGY')

    def table_has_ontology_types(self, table_meta: Table) -> bool:
        """Determine if there are ONTOLOGY types in a table"""
        count: int = 0
        for column in table_meta:
            if self.column_is_ontology_type(column=column):
                count += 1
        return count > 0

    def build_sheet(self,
                    workbook, sheet_name: str,
                    column_metadata: list[Column],
                    styles: WorkbookStyles):
        """Build worksheet from schema metadata

        :param sheet_name: name of the new workbook sheet to create
        :type sheet_name: str

        :param column_metadata: column metadata from emx2 pyclient
        :type column_metadata: list[Column]
        """
        new_sheet = workbook.add_worksheet(name=sheet_name)
        index: int = 0
        for column in column_metadata:
            log.info('Processing column %s', column.name)
            # write header
            self.write_sheet_header(sheet=new_sheet,
                                    column=column,
                                    styles=styles,
                                    col_index=index)

            # determine if ontology table is present
            ontology_table: str = column.get('refTableName')
            should_build_ontology: bool = self.column_is_ontology_type(
                column=column) and ontology_table is not None

            if should_build_ontology:
                log.info('Creating lookup from %s', ontology_table)
                self.should_build_lookup_sheet = True
                ontology_schema: str = self.schema

                if bool(column.get('refSchemaId')):
                    ontology_schema = column.refSchemaId

                data = client.get(table=ontology_table,
                                  columns=['name'],
                                  schema=ontology_schema)

                lookups_col: str = get_column_letter(self.lookups_col_index+1)
                lookup = {
                    'name': ontology_table,
                    'data': list(data)[:10],
                    'lookups_col': lookups_col,
                    'lookups_col_index': self.lookups_col_index,
                    'template_sheet': sheet_name,
                    'template_col': get_column_letter(index+1),
                    'template_col_index': index,
                    'formula': f"=lookups!{lookups_col}2:{lookups_col}{len(data)}"
                }
                # print(lookup)
                self.lookups.append(lookup)
                self.lookups_col_index += 1

            # iterate over rows in the sheet: apply styles and/or validation
            if self.column_is_required(column=column):
                for row_index in range(1, self.max_template_rows):
                    new_sheet.write(row_index,
                                    index,
                                    None,
                                    styles['cell_required'])

            index += 1
        new_sheet.autofit()

    def build(self, metadata: Schema):
        """Build template"""
        workbook = xlsxwriter.Workbook(filename=self.output_filename)

        # set styles
        styles: WorkbookStyles = {
            'header_default': workbook.add_format({'border': 1}),
            'header_required': workbook.add_format({
                'bottom': 1,
                'bg_color': '#ADE1FF',
                'bold': True
            }),
            'cell_required': workbook.add_format({
                'border': 1,
                'bg_color': '#cbcbcb'
            })
        }
        styles['cell_required'].set_border_color('#cbcbcb')

        for table in self.tables:
            log.info('Building sheet for %s', table)
            table_meta = metadata.get_table(by='name', value=table)

            excluded_types = ['SECTION', 'HEADING']
            col_meta = [
                col for col in table_meta.columns
                if col.columnType not in excluded_types and not col.name.startswith('mg_')
            ]

            self.build_sheet(workbook=workbook,
                             sheet_name=table,
                             column_metadata=col_meta,
                             styles=styles)

        # only build lookups if present in the model
        if self.should_build_lookup_sheet:
            lookups_sheet = workbook.add_worksheet(name='lookups')
            for lookup in self.lookups:
                log.info(
                    'Creating lookup and apply validation rules for %s', lookup['name'])
                lookups_sheet.write(
                    0,
                    lookup['lookups_col_index'],
                    lookup['name'],
                    styles['header_default'])

                # write ontology terms
                for index, row in enumerate(lookup['data']):
                    lookups_sheet.write(
                        index+1, lookup['lookups_col_index'], f"'{row['name']}'")

                # apply validation in the appropriate sheet
                template_sheet = workbook.get_worksheet_by_name(
                    lookup['template_sheet'])
                if lookup['data']:
                    for row_index in range(1, self.max_template_rows):
                        template_sheet.data_validation(
                            f"{lookup['template_col']}{row_index+1}",
                            {'validate': 'list',
                                'source': f"{lookup['formula']}"}
                        )

            lookups_sheet.autofit()
            lookups_sheet.protect()
        workbook.close()


if __name__ == "__main__":
    log.info("Building bulk upload template from %s", SCHEMA)

    schema_meta = client.get_schema_metadata(name=SCHEMA)
    template = BuildTemplate(
        schema=SCHEMA,
        tables=TABLES,
        max_template_rows=MAX_TEMPLATE_ROWS,
        sys_output_filename=OUTPUT_FILE
    )
    template.build(metadata=schema_meta)
